"""
Excel Export Service
Generates formatted Excel reports from MongoDB review data.
"""
import io
from datetime import datetime, timedelta
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import structlog
from app.core.database import get_collection

log = structlog.get_logger()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
POS_FILL = PatternFill("solid", fgColor="E2EFDA")
NEG_FILL = PatternFill("solid", fgColor="FCE4D6")


class ExcelExportService:

    def _collection(self):
        return get_collection("reviews")

    async def generate_report(
        self,
        tenant_id: UUID,
        location_id: UUID | None = None,
        platform: str | None = None,
        days: int = 30,
    ) -> bytes:
        """
        Generate a full Excel report with 3 sheets:
        1. Summary — overall stats
        2. Reviews — all reviews with sentiment
        3. Trends — daily breakdown
        """
        since = datetime.utcnow() - timedelta(days=days)
        match = {"tenant_id": str(tenant_id), "published_at": {"$gte": since}}
        if location_id:
            match["location_id"] = str(location_id)
        if platform:
            match["platform"] = platform

        collection = self._collection()
        reviews = await collection.find(match, {"_id": 0}).sort("published_at", -1).to_list(length=5000)

        wb = openpyxl.Workbook()

        self._build_summary_sheet(wb, reviews, days)
        self._build_reviews_sheet(wb, reviews)
        self._build_trends_sheet(wb, reviews)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _build_summary_sheet(self, wb, reviews, days):
        ws = wb.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        total = len(reviews)
        avg_rating = sum(r.get("rating", 0) for r in reviews) / total if total else 0
        positive = sum(1 for r in reviews if r.get("sentiment", {}).get("label") == "positive")
        negative = sum(1 for r in reviews if r.get("sentiment", {}).get("label") == "negative")
        neutral = sum(1 for r in reviews if r.get("sentiment", {}).get("label") == "neutral")

        # Title
        ws["A1"] = "Reputation Intelligence Report"
        ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
        ws["A3"] = f"Period: Last {days} days"

        rows = [
            ("", ""),
            ("METRIC", "VALUE"),
            ("Total Reviews", total),
            ("Average Rating", round(avg_rating, 2)),
            ("Positive Reviews", positive),
            ("Negative Reviews", negative),
            ("Neutral Reviews", neutral),
            ("Positive %", f"{round(positive/total*100, 1)}%" if total else "0%"),
            ("Negative %", f"{round(negative/total*100, 1)}%" if total else "0%"),
        ]

        for i, (label, value) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            if label == "METRIC":
                for col in [1, 2]:
                    cell = ws.cell(row=i, column=col)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

    def _build_reviews_sheet(self, wb, reviews):
        ws = wb.create_sheet("Reviews")

        headers = [
            "Date", "Platform", "Author", "Rating",
            "Sentiment", "Score", "Content", "Owner Reply"
        ]
        col_widths = [15, 18, 20, 10, 12, 10, 60, 40]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = width

        for row_idx, review in enumerate(reviews, start=2):
            published = review.get("published_at", "")
            if isinstance(published, datetime):
                published = published.strftime("%Y-%m-%d")

            sentiment = review.get("sentiment", {})
            label = sentiment.get("label", "")
            score = sentiment.get("score", 0)

            values = [
                published,
                review.get("platform", ""),
                review.get("author", {}).get("name", ""),
                review.get("rating", ""),
                label,
                score,
                review.get("content", ""),
                review.get("owner_reply", ""),
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color code by sentiment
            fill = POS_FILL if label == "positive" else NEG_FILL if label == "negative" else None
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    def _build_trends_sheet(self, wb, reviews):
        ws = wb.create_sheet("Daily Trends")

        headers = ["Date", "Review Count", "Avg Rating", "Positive", "Negative", "Neutral"]
        col_widths = [15, 15, 15, 12, 12, 12]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.column_dimensions[get_column_letter(col)].width = width

        # Group by date
        daily: dict[str, dict] = {}
        for review in reviews:
            published = review.get("published_at", "")
            if isinstance(published, datetime):
                date_key = published.strftime("%Y-%m-%d")
            else:
                date_key = str(published)[:10]

            if date_key not in daily:
                daily[date_key] = {"count": 0, "ratings": [], "positive": 0, "negative": 0, "neutral": 0}

            daily[date_key]["count"] += 1
            daily[date_key]["ratings"].append(review.get("rating", 0))
            label = review.get("sentiment", {}).get("label", "neutral")
            if label in daily[date_key]:
                daily[date_key][label] += 1

        for row_idx, (date, data) in enumerate(sorted(daily.items(), reverse=True), start=2):
            avg = sum(data["ratings"]) / len(data["ratings"]) if data["ratings"] else 0
            values = [date, data["count"], round(avg, 2), data["positive"], data["negative"], data["neutral"]]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=value)

        ws.freeze_panes = "A2"


excel_export_service = ExcelExportService()
